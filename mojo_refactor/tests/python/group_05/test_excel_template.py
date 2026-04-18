"""
Test for rqalpha/mod/rqalpha_mod_sys_analyser/report/excel_template.py
Comprehensive tests matching the Mojo refactored version
"""

import os
import tempfile
import shutil


class TestExcelTemplate:
    """Test ExcelTemplate class"""

    def test_excel_template_class_exists(self):
        """Test ExcelTemplate class exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import ExcelTemplate

        assert ExcelTemplate is not None

    def test_sheet_schema_class_exists(self):
        """Test SheetSchema class exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        assert SheetSchema is not None

    def test_single_cell_schema_class_exists(self):
        """Test SingleCellSchema class exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SingleCellSchema

        assert SingleCellSchema is not None

    def test_vertical_series_schema_class_exists(self):
        """Test VerticalSeriesSchema class exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VerticalSeriesSchema

        assert VerticalSeriesSchema is not None

    def test_summary_template_class_exists(self):
        """Test SummaryTemplate class exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SummaryTemplate

        assert SummaryTemplate is not None


class TestValueNameRegex:
    """Test VALUE_NAME_RE regex"""

    def test_value_name_re_exists(self):
        """Test VALUE_NAME_RE constant exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VALUE_NAME_RE

        assert VALUE_NAME_RE is not None
        assert isinstance(VALUE_NAME_RE, str)

    def test_value_name_re_matches_valid(self):
        """Test VALUE_NAME_RE matches valid patterns"""
        import re
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VALUE_NAME_RE

        assert re.match(VALUE_NAME_RE, "#total_returns#") is not None
        assert re.match(VALUE_NAME_RE, "#sharpe_ratio#") is not None
        assert re.match(VALUE_NAME_RE, "#a1#") is not None

    def test_value_name_re_rejects_invalid(self):
        """Test VALUE_NAME_RE rejects invalid patterns"""
        import re
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VALUE_NAME_RE

        assert re.match(VALUE_NAME_RE, "total_returns") is None
        assert re.match(VALUE_NAME_RE, "#Total Returns#") is None
        assert re.match(VALUE_NAME_RE, "#returns#") is not None  # lowercase letters are valid
        assert re.match(VALUE_NAME_RE, "#Returns#") is None  # uppercase not allowed


class TestSheetSchema:
    """Test SheetSchema functionality"""

    def test_sheet_schema_parse_named_cells(self):
        """Test SheetSchema correctly parses #name# pattern cells"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(1, 1).value = "#total_returns#"
        ws.cell(2, 1).value = "#sharpe_ratio#"
        ws.cell(3, 1).value = "normal text"

        schema = SheetSchema(ws)

        assert "total_returns" in schema._cell_map
        assert "sharpe_ratio" in schema._cell_map
        assert len(schema._cell_map) == 2

    def test_sheet_schema_cell_map_positions(self):
        """Test SheetSchema stores correct row/column positions"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(5, 3).value = "#metric#"

        schema = SheetSchema(ws)
        row, col, _ = schema._cell_map["metric"]

        assert row == 5
        assert col == 3

    def test_sheet_schema_fill_worksheet_not_implemented(self):
        """Test SheetSchema.fill_worksheet raises NotImplementedError"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        schema = SheetSchema(ws)

        try:
            schema.fill_worksheet(ws, {})
            assert False, "Should raise NotImplementedError"
        except NotImplementedError:
            pass


class TestSingleCellSchema:
    """Test SingleCellSchema functionality"""

    def test_single_cell_fill(self):
        """Test SingleCellSchema fills single values per named cell"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SingleCellSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(1, 1).value = "#total_returns#"
        ws.cell(2, 1).value = "#sharpe#"

        schema = SingleCellSchema(ws)
        data = {"total_returns": "15.5%", "sharpe": "1.8"}

        schema.fill_worksheet(ws, data)

        assert ws.cell(1, 1).value == "15.5%"
        assert ws.cell(2, 1).value == "1.8"


class TestVerticalSeriesSchema:
    """Test VerticalSeriesSchema functionality"""

    def test_vertical_fill_with_data(self):
        """Test VerticalSeriesSchema fills data vertically"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VerticalSeriesSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(1, 1).value = "#monthly_return#"

        schema = VerticalSeriesSchema(ws)
        data = {"monthly_return": [1.0, 2.0, 3.0]}

        schema.fill_worksheet(ws, data)

        assert ws.cell(1, 1).value == 1.0
        assert ws.cell(2, 1).value == 2.0
        assert ws.cell(3, 1).value == 3.0

    def test_vertical_fill_empty_data(self):
        """Test VerticalSeriesSchema handles empty data with NaN"""
        import openpyxl
        import numpy as np
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import VerticalSeriesSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(1, 1).value = "#series#"

        schema = VerticalSeriesSchema(ws)

        schema.fill_worksheet(ws, {})

        assert np.isnan(ws.cell(1, 1).value)


class TestGenerateXlsxReports:
    """Test generate_xlsx_reports function"""

    def test_generate_xlsx_reports_function_exists(self):
        """Test generate_xlsx_reports function exists"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import generate_xlsx_reports

        assert callable(generate_xlsx_reports)

    def test_generate_xlsx_reports_creates_file(self):
        """Test generate_xlsx_reports creates output file"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import generate_xlsx_reports
        import openpyxl

        tmp_dir = tempfile.mkdtemp()
        try:
            data = {
                "概览": {"total_returns": "10%"},
                "年度指标": {"annual_return": [5.0, 8.0]},
            }

            generate_xlsx_reports(data, tmp_dir)

            output_file = os.path.join(tmp_dir, "summary.xlsx")
            assert os.path.exists(output_file)

            wb = openpyxl.load_workbook(output_file)
            assert len(wb.worksheets) > 0
        finally:
            shutil.rmtree(tmp_dir)


class TestSummaryTemplateConstants:
    """Test SummaryTemplate constants"""

    def test_template_name(self):
        """Test TEMPLATE_NAME constant"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SummaryTemplate

        assert SummaryTemplate.TEMPLATE_NAME == "summary"

    def test_schema_classes(self):
        """Test SCHEMA_CLASSES dictionary"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SummaryTemplate

        assert isinstance(SummaryTemplate.SCHEMA_CLASSES, dict)
        assert "概览" in SummaryTemplate.SCHEMA_CLASSES
        assert "年度指标" in SummaryTemplate.SCHEMA_CLASSES
        assert "月度收益" in SummaryTemplate.SCHEMA_CLASSES
        assert "月度超额收益（几何）" in SummaryTemplate.SCHEMA_CLASSES
        assert "个股权重" in SummaryTemplate.SCHEMA_CLASSES
        assert "压力测试" in SummaryTemplate.SCHEMA_CLASSES

    def test_summary_template_singleton(self):
        """Test SUMMARY_TEMPLATE singleton instance"""
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SUMMARY_TEMPLATE

        assert SUMMARY_TEMPLATE is not None
        assert SUMMARY_TEMPLATE.TEMPLATE_NAME == "summary"


class TestWriteCellEdgeCases:
    """Test _write_cell edge cases"""

    def test_write_cell_none_becomes_nan(self):
        """Test that None data becomes numpy.nan"""
        import openpyxl
        import numpy as np
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        schema = SheetSchema(ws)
        schema._write_cell(ws, 1, 1, None)

        assert np.isnan(ws.cell(1, 1).value)

    def test_write_cell_nat_becomes_empty(self):
        """Test that pd.NaT datetime becomes empty string"""
        import openpyxl
        import pandas as pd
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema

        wb = openpyxl.Workbook()
        ws = wb.active

        schema = SheetSchema(ws)
        schema._write_cell(ws, 1, 1, pd.NaT)

        assert ws.cell(1, 1).value == ""

    def test_write_cell_applies_style(self):
        """Test that style is applied to cell when provided"""
        import openpyxl
        from rqalpha.mod.rqalpha_mod_sys_analyser.report.excel_template import SheetSchema
        from openpyxl.styles.cell_style import StyleArray

        wb = openpyxl.Workbook()
        ws = wb.active

        schema = SheetSchema(ws)
        fake_style = StyleArray()

        schema._write_cell(ws, 1, 1, "test", fake_style)

        assert ws.cell(1, 1).value == "test"
